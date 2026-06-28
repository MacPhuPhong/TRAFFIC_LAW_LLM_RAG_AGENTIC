# -*- coding: utf-8 -*-
"""Export 25-question evaluation set to Excel.

Output: research/results/eval_25_questions.xlsx with two sheets:
  1. "Summary" — one row per question × pipeline, side-by-side comparison.
  2. "Full Detail" — one row per (question, pipeline) with full answer, citations, contexts count, latency, metrics.
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent.parent  # traffic_rag/
EVAL = ROOT / "research" / "data" / "eval_qa.jsonl"
METRICS = ROOT / "research" / "results" / "metrics"
OUT = ROOT / "research" / "results" / "eval_25_questions.xlsx"

PIPELINES = [
    ("gemini_only", "Gemini Only"),
    ("vanilla_rag", "Vanilla RAG"),
    ("agentic_rag", "Agentic RAG"),
]


def load_jsonl(p: Path) -> list[dict]:
    return [json.loads(l) for l in open(p, encoding="utf-8") if l.strip()]


def fmt_citation(c: dict) -> str:
    parts = []
    if c.get("dieu") is not None and c.get("dieu") != 0:
        parts.append(f"Điều {c['dieu']}")
    if c.get("khoan"):
        parts.append(f"Khoản {c['khoan']}")
    if c.get("diem"):
        parts.append(f"Điểm {c['diem']}")
    doc = c.get("doc_id", "")
    head = " · ".join(parts)
    return f"[{head} — {doc}]" if head else f"[{doc}]"


def fmt_citations(cites: list[dict]) -> str:
    if not cites:
        return ""
    return "\n".join(fmt_citation(c) for c in cites)


def cite_match(pred: list[dict], gold: list[dict]) -> str:
    """Return human label: ✅ exact / ⚠ doc-only / ❌ none / — out_of_scope."""
    if not gold:
        return "— (out_of_scope)"
    gold_keys_khoan = {(g.get("doc_id"), g.get("dieu"), str(g.get("khoan") or "")) for g in gold}
    gold_keys_doc = {g.get("doc_id") for g in gold}
    pred_keys_khoan = {(p.get("doc_id"), p.get("dieu"), str(p.get("khoan") or "")) for p in pred}
    pred_keys_doc = {p.get("doc_id") for p in pred}
    if gold_keys_khoan & pred_keys_khoan:
        return "✅ Khớp khoản"
    if gold_keys_doc & pred_keys_doc:
        return "⚠ Khớp văn bản"
    if not pred:
        return "❌ Không trích"
    return "❌ Sai văn bản"


def main():
    gold_rows = load_jsonl(EVAL)
    gold_by_id = {r["id"]: r for r in gold_rows}

    pipelines_data = {}
    for name, _ in PIPELINES:
        path = METRICS / f"rq1_{name}.jsonl"
        records = load_jsonl(path)
        pipelines_data[name] = {r["id"]: r for r in records}

    # === Sheet 1: Summary side-by-side ===
    summary_rows = []
    for g in gold_rows:
        qid = g["id"]
        row = {
            "ID": qid,
            "Category": g.get("category", ""),
            "Câu hỏi": g["question"],
            "Gold answer": g.get("gold_answer", ""),
            "Gold citations": fmt_citations(g.get("gold_citations", [])),
        }
        for name, label in PIPELINES:
            rec = pipelines_data[name].get(qid, {})
            row[f"{label} — Answer"] = rec.get("answer", "")
            row[f"{label} — Citations"] = fmt_citations(rec.get("citations", []))
            row[f"{label} — Cit-Match"] = cite_match(
                rec.get("citations", []),
                g.get("gold_citations", []),
            )
            row[f"{label} — Latency (s)"] = round(rec.get("latency_s", 0), 2)
        summary_rows.append(row)
    summary_df = pd.DataFrame(summary_rows)

    # === Sheet 2: Full Detail ===
    detail_rows = []
    for g in gold_rows:
        qid = g["id"]
        for name, label in PIPELINES:
            rec = pipelines_data[name].get(qid, {})
            detail_rows.append({
                "ID": qid,
                "Category": g.get("category", ""),
                "Pipeline": label,
                "Câu hỏi": g["question"],
                "Gold answer": g.get("gold_answer", ""),
                "Gold citations": fmt_citations(g.get("gold_citations", [])),
                "Expected category": g.get("expected_category", ""),
                "Predicted answer": rec.get("answer", ""),
                "Predicted citations": fmt_citations(rec.get("citations", [])),
                "Cit-match label": cite_match(rec.get("citations", []), g.get("gold_citations", [])),
                "Predicted category (router)": rec.get("category", ""),
                "Contexts retrieved (count)": len(rec.get("contexts", []) or []),
                "Latency (s)": round(rec.get("latency_s", 0), 2),
                "Error": rec.get("error") or "",
            })
    detail_df = pd.DataFrame(detail_rows)

    # === Sheet 3: Pipeline summary (mean metrics) ===
    rq1 = pd.read_csv(METRICS / "rq1_summary.csv")
    pipeline_summary = rq1.copy()
    pipeline_summary.columns = ["Pipeline", "F1 (token)", "ROUGE-L", "Mean latency (s)"]
    pipeline_summary["Pipeline"] = pipeline_summary["Pipeline"].map({
        "gemini_only": "Gemini Only",
        "vanilla_rag": "Vanilla RAG",
        "agentic_rag": "Agentic RAG",
    }).fillna(pipeline_summary["Pipeline"])

    # === Write workbook ===
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Side-by-side (25 câu)", index=False)
        detail_df.to_excel(writer, sheet_name="Full Detail (75 dòng)", index=False)
        pipeline_summary.to_excel(writer, sheet_name="Pipeline Summary", index=False)

    # === Post-format ===
    from openpyxl import load_workbook
    wb = load_workbook(OUT)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    col_widths = {
        "Side-by-side (25 câu)": {
            "A": 8,  "B": 16, "C": 50, "D": 60, "E": 40,
            "F": 60, "G": 40, "H": 22, "I": 12,
            "J": 60, "K": 40, "L": 22, "M": 12,
            "N": 60, "O": 40, "P": 22, "Q": 12,
        },
        "Full Detail (75 dòng)": {
            "A": 8, "B": 16, "C": 14, "D": 50, "E": 60, "F": 40,
            "G": 16, "H": 70, "I": 40, "J": 22, "K": 18, "L": 16, "M": 12, "N": 20,
        },
        "Pipeline Summary": {"A": 16, "B": 14, "C": 14, "D": 18},
    }

    for sheet_name, widths in col_widths.items():
        ws = wb[sheet_name]
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        # header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30
        # body
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border
        # data row height
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 120
        ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"  - Sheet 1: {len(summary_df)} rows (1 row/question, 3 pipelines side-by-side)")
    print(f"  - Sheet 2: {len(detail_df)} rows (1 row/question/pipeline)")
    print(f"  - Sheet 3: {len(pipeline_summary)} rows (aggregate metrics)")


if __name__ == "__main__":
    main()
